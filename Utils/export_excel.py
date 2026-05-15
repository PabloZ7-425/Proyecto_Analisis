# utils/export_excel.py
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

class ExportadorExcel:
    def __init__(self, datos_ventas_mes, datos_diarios, nombre_mes):
        """
        datos_ventas_mes: lista de todas las ventas del mes (cada una debe tener 'fecha_hora', 'numero_documento', etc.)
        datos_diarios: dict {fecha_str: lista_ventas_de_ese_dia}
        nombre_mes: string ej. "ENERO"
        """
        self.ventas_mes = datos_ventas_mes
        self.diarios = datos_diarios
        self.nombre_mes = nombre_mes.upper()

    def generar(self, nombre_archivo=None):
        if not nombre_archivo:
            nombre_archivo = f"Reporte_{self.nombre_mes}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Hoja resumen (mes)
        self._crear_hoja_resumen(wb)
        # Hojas diarias
        for fecha_str, ventas_dia in self.diarios.items():
            self._crear_hoja_diaria(wb, fecha_str, ventas_dia)

        wb.save(nombre_archivo)
        return nombre_archivo

    def _crear_hoja_resumen(self, wb):
        ws = wb.create_sheet(self.nombre_mes)
        # Estilos
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="F5C800", end_color="F5C800", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        headers = ["DIA", "FAC./RECI", "NO", "CLIENTE", "PRODUCTO", "Marca", "MODELO",
                   "SALE DE BODEGA 1 GUATE", "SALE DE BODEGA 2 XELA", "FORMA PAGO",
                   "CONSIGNATARIO", "SOPORTE", "BODEGA DE ENTREGA"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Llenar filas
        for i, venta in enumerate(self.ventas_mes, start=2):
            fecha = venta.get('fecha_hora')
            dia = fecha.day if fecha else ''
            ws.cell(row=i, column=1, value=dia)
            ws.cell(row=i, column=2, value=venta.get('numero_documento', ''))
            ws.cell(row=i, column=3, value=venta.get('id_venta', ''))
            cliente = f"{venta.get('cliente_nombre','')} {venta.get('cliente_apellido','')}".strip()
            ws.cell(row=i, column=4, value=cliente)
            # Tomamos el primer producto (puedes mejorar si hay múltiples)
            productos = venta.get('productos', [])
            if productos:
                ws.cell(row=i, column=5, value=productos[0].get('producto_nombre', ''))
                ws.cell(row=i, column=6, value=productos[0].get('marca', ''))
                ws.cell(row=i, column=7, value=productos[0].get('modelo', ''))
            else:
                ws.cell(row=i, column=5, value='')
                ws.cell(row=i, column=6, value='')
                ws.cell(row=i, column=7, value='')
            # Bodega 1 (Guate) siempre 0
            ws.cell(row=i, column=8, value=0)
            # Bodega 2 (XELA) = total de la venta
            ws.cell(row=i, column=9, value=float(venta.get('total', 0)))
            ws.cell(row=i, column=10, value=venta.get('forma_pago', ''))
            # Consignatario: si es envío, la empresa; si no, "TX"
            consignatario = venta.get('empresa_envio', '') if venta.get('es_envio') else 'TX'
            ws.cell(row=i, column=11, value=consignatario)
            ws.cell(row=i, column=12, value='')  # Soporte vacío
            # Bodega de entrega: "XELA" por defecto (aunque envíos podrían ser "GUATE", tú decides)
            entrega = 'XELA'
            ws.cell(row=i, column=13, value=entrega)

        # Ajustar anchos
        for col in range(1, len(headers)+1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _crear_hoja_diaria(self, wb, fecha_str, ventas_dia):
        # fecha_str ejemplo "2026-01-15"
        dia = fecha_str.split('-')[2]
        ws = wb.create_sheet(dia)

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        # Títulos
        ws['A1'] = "TEC-SHOP"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A3'] = "CONTROL DE MOVIMIENTOS DIARIOS"
        ws['A5'] = "FECHA:"
        ws['B5'] = fecha_str

        # Denominaciones (puedes dejarlo en cero o cargar desde detalle_apertura si quieres)
        ws['A7'] = "DENOMINACION"
        ws['B7'] = "CANTIDAD"
        ws['C7'] = "MONTO"
        denoms = [100, 50, 20, 10, 5, 1]
        for idx, d in enumerate(denoms, start=8):
            ws.cell(row=idx, column=1, value=d)
            ws.cell(row=idx, column=2, value=0)
            ws.cell(row=idx, column=3, value=f"=B{idx}*A{idx}")
        ws['D16'] = "TOTAL EFECTIVO"
        ws['E16'] = f"=SUM(C8:C{len(denoms)+7})"

        # Ingresos del día
        ws['G5'] = "INGRESOS"
        ws['G7'] = "FECHA"
        ws['H7'] = "DOCTO"
        ws['I7'] = "MONTO"
        ws['J7'] = "FORMA"
        row = 8
        for venta in ventas_dia:
            fecha = venta.get('fecha_hora')
            ws.cell(row=row, column=7, value=fecha.strftime('%Y-%m-%d') if fecha else '')
            ws.cell(row=row, column=8, value=venta.get('numero_documento', ''))
            ws.cell(row=row, column=9, value=float(venta.get('total', 0)))
            ws.cell(row=row, column=10, value=venta.get('forma_pago', ''))
            row += 1
        ws['I19'] = "TOTAL EFECTIVO"
        ws['J19'] = f"=SUM(I8:I{row-1})"

        # Sección de cuentas por cobrar (opcional)
        ws['A22'] = "CUENTAS POR COBRAR"
        ws['A24'] = "FECHA"
        ws['B24'] = "No. Dcto"
        ws['C24'] = "MONTO"
        ws['D24'] = "GUIA"
        # Puedes llenar aquí las cuentas pendientes si lo deseas

        # Totales finales (como en tu Excel)
        ws['A31'] = "VENTAS DE XELA -BODEGA XELA"
        ws['F31'] = "VENTAS DE XELA"
        ws['F32'] = f"=SUM(I8:I{row-1})"

        # Aplicar bordes
        for r in range(7, row):
            for c in [7,8,9,10]:
                ws.cell(row=r, column=c).border = thin_border